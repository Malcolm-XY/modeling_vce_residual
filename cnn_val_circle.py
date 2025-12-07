# -*- coding: utf-8 -*-
"""
Created on Sun Dec 22 01:23:16 2024

@author: 18307
"""
import os
import numpy as np
import pandas as pd

# %% end program
import time
import threading
def shutdown_with_countdown(countdown_seconds=30):
    """
    Initiates a shutdown countdown, allowing the user to cancel shutdown within the given time.

    Args:
        countdown_seconds (int): The number of seconds to wait before shutting down.
    """
    def cancel_shutdown():
        nonlocal shutdown_flag
        user_input = input("\nPress 'c' and Enter to cancel shutdown: ").strip().lower()
        if user_input == 'c':
            shutdown_flag = False
            print("Shutdown cancelled.")

    # Flag to determine whether to proceed with shutdown
    shutdown_flag = True

    # Start a thread to listen for user input
    input_thread = threading.Thread(target=cancel_shutdown, daemon=True)
    input_thread.start()

    # Countdown timer
    print(f"Shutdown scheduled in {countdown_seconds} seconds. Press 'c' to cancel.")
    for i in range(countdown_seconds, 0, -1):
        print(f"Time remaining: {i} seconds", end="\r")
        time.sleep(1)

    # Check the flag after countdown
    if shutdown_flag:
        print("\nShutdown proceeding...")
        os.system("shutdown /s /t 1")  # Execute shutdown command
    else:
        print("\nShutdown aborted.")

def end_program_actions(play_sound=True, shutdown=False, countdown_seconds=120):
    """
    Performs actions at the end of the program, such as playing a sound or shutting down the system.

    Args:
        play_sound (bool): If True, plays a notification sound.
        shutdown (bool): If True, initiates shutdown with a countdown.
        countdown_seconds (int): Countdown time for shutdown confirmation.
    """
    if play_sound:
        try:
            import winsound
            print("Playing notification sound...")
            winsound.Beep(1000, 500)  # Frequency: 1000Hz, Duration: 500ms
        except ImportError:
            print("winsound module not available. Skipping sound playback.")

    if shutdown:
        shutdown_with_countdown(countdown_seconds)

# %% read parameters/save
def read_params(model='exponential', model_fm='basic', model_rcm='differ', folder='fitting_results(15_15_joint_band_from_mat)'):
    identifier = f'{model_fm.lower()}_fm_{model_rcm.lower()}_rcm'
    
    path_current = os.getcwd()
    path_fitting_results = os.path.join(path_current, 'fitting_results', folder)
    file_path = os.path.join(path_fitting_results, f'fitting_results({identifier}).xlsx')
    
    df = pd.read_excel(file_path).set_index('method')
    df_dict = df.to_dict(orient='index')
    
    model = model.upper()
    params = df_dict[model]
    
    return params

from openpyxl import load_workbook
def save_results_to_xlsx_append(results, output_dir, filename, sheet_name='K-Fold Results'):
    """
    Appends results to an existing Excel file or creates a new file if it doesn't exist.

    Args:
        results (list or pd.DataFrame): The results data to save.
        output_dir (str): The directory where the Excel file will be saved.
        filename (str): The name of the Excel file.
        sheet_name (str): The sheet name in the Excel file. Default is 'K-Fold Results'.

    Returns:
        str: The path of the saved Excel file.
    """
    # Convert results to DataFrame if necessary
    if not isinstance(results, pd.DataFrame):
        results_df = pd.DataFrame(results)
    else:
        results_df = results

    # Rearrange columns if "Identifier" is present
    if 'Identifier' in results_df.columns:
        columns_order = ['Identifier'] + [col for col in results_df.columns if col != 'Identifier']
        results_df = results_df[columns_order]

    # Ensure output directory exists
    os.makedirs(output_dir, exist_ok=True)

    # Define the full output path
    output_path = os.path.join(output_dir, filename)

    # Append to existing Excel file or create a new one
    if os.path.exists(output_path):
        print(f"Appending data to existing file: {output_path}")
        with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # Get the existing workbook
            existing_workbook = load_workbook(output_path)

            # Check if the sheet exists
            if sheet_name in existing_workbook.sheetnames:
                # Load existing sheet and append
                start_row = existing_workbook[sheet_name].max_row
                results_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=start_row)
            else:
                # Write new sheet if not exists
                results_df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        print(f"Creating new file: {output_path}")
        results_df.to_excel(output_path, index=False, sheet_name=sheet_name)

    print(f"Results successfully saved to: {output_path}")
    return output_path

# %% control; original cm
import torch
import feature_engineering
import cnn_validation
from models import models

def cnn_evaluation_circle_original_cm(feature_cm, subject_range=range(1, 6), experiment_range=range(1, 4), save=False):
    # labels
    labels = utils_feature_loading.read_labels(dataset='seed')
    y = torch.tensor(np.array(labels)).view(-1)
    
    # data and evaluation circle
    all_results_original = []
    for sub in subject_range:
        for ex in experiment_range:
            subject_id = f"sub{sub}ex{ex}"
            print(f"Evaluating {subject_id}...")
            
            # CM/MAT
            features = utils_feature_loading.read_fcs_mat('seed', subject_id, feature_cm)
            alpha = features['alpha']
            beta = features['beta']
            gamma = features['gamma']
            
            # CM/H5
            # features = utils_feature_loading.read_fcs('seed', subject_id, feature_cm)
            # alpha = features['alpha']
            # beta = features['beta']
            # gamma = features['gamma']
            
            x = np.stack((alpha, beta, gamma), axis=1)
            
            # cnn model
            cnn_model = models.CNN_2layers_adaptive_maxpool_3()
            # traning and testing
            result_CM = cnn_validation.cnn_cross_validation(cnn_model, x, y)
            
            # Add identifier to the result
            result_CM['Identifier'] = f'sub{sub}ex{ex}'
            
            all_results_original.append(result_CM)
    
    # save
    output_dir = os.path.join(os.getcwd(), 'results_cnn_evaluation')
    filename_CM = f"cnn_validation_CM_{feature_cm}.xlsx"
    if save: save_results_to_xlsx_append(all_results_original, output_dir, filename_CM)
    
    return all_results_original

# %% experiment; rebuilded cm
from utils import utils_feature_loading
from connectivity_matrix_rebuilding import cm_rebuilding as cm_rebuild

def cnn_evaluation_circle_rebuilded_cm(feature_cm, model, model_fm, model_rcm, 
                                       subject_range=range(11, 16), experiment_range=range(1, 4), save=False):
    # labels
    labels = utils_feature_loading.read_labels(dataset='seed')
    y = torch.tensor(np.array(labels)).view(-1)
    
    # distance matrix
    _, dm = feature_engineering.compute_distance_matrix(dataset="seed", projection_params={"type": "3d"}, visualize=True)
    dm = feature_engineering.normalize_matrix(dm)
    
    # parameters for construction of FM and RCM
    param = read_params(model, model_fm, model_rcm, folder='fitting_results(10_15_joint_band_from_mat)')
    
    # data and evaluation circle
    all_results_rebuilded = []
    average_accuracy_rebuilded, average_accuracy_rebuilded_counter = 0.0, 0
    
    for sub in subject_range:
        for ex in experiment_range:
            subject_id = f"sub{sub}ex{ex}"
            print(f"Evaluating {subject_id}...")
            
            # CM/MAT
            features = utils_feature_loading.read_fcs_mat('seed', subject_id, feature_cm)
            alpha = features['alpha']
            beta = features['beta']
            gamma = features['gamma']

            # CM/H5
            # features = utils_feature_loading.read_fcs('seed', subject_id, feature_cm)
            # alpha = features['alpha']
            # beta = features['beta']
            # gamma = features['gamma']

            # RCM
            alpha_rebuilded = cm_rebuild(alpha, dm, param, model, model_fm, model_rcm, True, False)
            beta_rebuilded = cm_rebuild(beta, dm, param, model, model_fm, model_rcm, True, False)
            gamma_rebuilded = cm_rebuild(gamma, dm, param, model, model_fm, model_rcm, True, False)
            
            x_rebuilded = np.stack((alpha_rebuilded, beta_rebuilded, gamma_rebuilded), axis=1)
            
            # cnn model
            cnn_model = models.CNN_2layers_adaptive_maxpool_3()
            # traning and testing
            result_RCM = cnn_validation.cnn_cross_validation(cnn_model, x_rebuilded, y)
            
            # Add identifier to the result
            result_RCM['Identifier'] = f'sub{sub}ex{ex}'
            all_results_rebuilded.append(result_RCM)
            
            average_accuracy_rebuilded += result_RCM['accuracy']
            average_accuracy_rebuilded_counter += 1

    average_accuracy_rebuilded = {'accuracy': average_accuracy_rebuilded/average_accuracy_rebuilded_counter}
    all_results_rebuilded.append(average_accuracy_rebuilded)
    
    # print(f'Final Results: {results_entry}')
    print('K-Fold Validation compelete\n')
    
    # save
    output_dir = os.path.join(os.getcwd(), 'results_cnn_evaluation')
    
    identifier = f'{model_fm.lower()}_fm_{model_rcm.lower()}_rcm'
    filename_RCM = f"cnn_validation_RCM({identifier})_{model}_{feature_cm}.xlsx"
    
    if save: save_results_to_xlsx_append(all_results_rebuilded, output_dir, filename_RCM)
    
    return all_results_rebuilded
    
if __name__ == '__main__':
    results_cm = cnn_evaluation_circle_original_cm('plv', range(11, 16), save=True)
    
    # %% linear
    model, model_fm, model_rcm = 'exponential', 'basic', 'linear'
    results_rcm = cnn_evaluation_circle_rebuilded_cm('plv', model, model_fm, model_rcm, 
                                           subject_range=range(11, 16), save=True)
    
    model, model_fm, model_rcm = 'gaussian', 'basic', 'linear'
    results_rcm = cnn_evaluation_circle_rebuilded_cm('plv', model, model_fm, model_rcm, 
                                           subject_range=range(11, 16), save=True)
    
    model, model_fm, model_rcm = 'generalized_gaussian', 'basic', 'linear'
    results_rcm = cnn_evaluation_circle_rebuilded_cm('plv', model, model_fm, model_rcm, 
                                           subject_range=range(11, 16), save=True)
    
    model, model_fm, model_rcm = 'powerlaw', 'basic', 'linear'
    results_rcm = cnn_evaluation_circle_rebuilded_cm('plv', model, model_fm, model_rcm, 
                                           subject_range=range(11, 16), save=True)
    
    model, model_fm, model_rcm = 'sigmoid', 'basic', 'linear'
    results_rcm = cnn_evaluation_circle_rebuilded_cm('plv', model, model_fm, model_rcm, 
                                           subject_range=range(11, 16), save=True)
    
    model, model_fm, model_rcm = 'inverse', 'basic', 'linear'
    results_rcm = cnn_evaluation_circle_rebuilded_cm('plv', model, model_fm, model_rcm, 
                                           subject_range=range(11, 16), save=True)
    
    model, model_fm, model_rcm = 'rational_quadratic', 'basic', 'linear'
    results_rcm = cnn_evaluation_circle_rebuilded_cm('plv', model, model_fm, model_rcm, 
                                           subject_range=range(11, 16), save=True)
    
    # %% linear_ratio
    model, model_fm, model_rcm = 'exponential', 'basic', 'linear_ratio'
    results_rcm = cnn_evaluation_circle_rebuilded_cm('plv', model, model_fm, model_rcm, 
                                           subject_range=range(11, 16), save=True)
    
    model, model_fm, model_rcm = 'gaussian', 'basic', 'linear_ratio'
    results_rcm = cnn_evaluation_circle_rebuilded_cm('plv', model, model_fm, model_rcm, 
                                           subject_range=range(11, 16), save=True)
    
    model, model_fm, model_rcm = 'generalized_gaussian', 'basic', 'linear_ratio'
    results_rcm = cnn_evaluation_circle_rebuilded_cm('plv', model, model_fm, model_rcm, 
                                           subject_range=range(11, 16), save=True)
    
    model, model_fm, model_rcm = 'powerlaw', 'basic', 'linear_ratio'
    results_rcm = cnn_evaluation_circle_rebuilded_cm('plv', model, model_fm, model_rcm, 
                                           subject_range=range(11, 16), save=True)
    
    model, model_fm, model_rcm = 'sigmoid', 'basic', 'linear_ratio'
    results_rcm = cnn_evaluation_circle_rebuilded_cm('plv', model, model_fm, model_rcm, 
                                           subject_range=range(11, 16), save=True)
    
    model, model_fm, model_rcm = 'inverse', 'basic', 'linear_ratio'
    results_rcm = cnn_evaluation_circle_rebuilded_cm('plv', model, model_fm, model_rcm, 
                                           subject_range=range(11, 16), save=True)
    
    model, model_fm, model_rcm = 'rational_quadratic', 'basic', 'linear_ratio'
    results_rcm = cnn_evaluation_circle_rebuilded_cm('pcc', model, model_fm, model_rcm, 
                                           subject_range=range(11, 16), save=True)
    
    # %% End
    end_program_actions(play_sound=True, shutdown=True, countdown_seconds=120)