# -*- coding: utf-8 -*-
"""
Created on Thu May  8 00:25:33 2025

@author: 18307
"""

import os
import numpy as np
import pandas as pd

import torch

from utils import utils_feature_loading, utils_visualization
import cnn_validation

def cnn_cross_validation_circle(model, mapping_func, dataset, dist, resolution, interp, feature, subject_range, experiment_range):
    """"
    Currently only support SEED dataset.
    """
    labels = torch.tensor(np.array(utils_feature_loading.read_labels(dataset)))
    targets = labels.view(-1)
    distribution = utils_feature_loading.read_distribution(dataset, dist)
    
    results_entry = []
    for sub in subject_range:
        for ex in experiment_range:
            identifier = f'sub{sub}ex{ex}'
            print(f'Processing {identifier}...')

            cfs = utils_feature_loading.read_cfs(dataset, identifier, feature)
            cfs_alpha = cfs['alpha']
            cfs_beta = cfs['beta']
            cfs_gamma = cfs['gamma']


            alpha_mapped = mapping_func(cfs_alpha, distribution, resolution, interpolation=interp)
            beta_mapped = mapping_func(cfs_beta, distribution, resolution, interpolation=interp)
            gamma_mapped = mapping_func(cfs_gamma, distribution, resolution, interpolation=interp, imshow=True)
            
            data_mapped = np.stack((alpha_mapped, beta_mapped, gamma_mapped), axis=1)
            
            # Training and Validation
            # Run validation in no_grad mode to avoid unnecessary gradient computation
            result = cnn_validation.cnn_cross_validation(model, data_mapped, targets)
            
            # Add identifier to the result
            result['Identifier'] = f'sub{sub}ex{ex}'
            results_entry.append(result)
            
    # print(f'Final Results: {results_entry}')
    print('K-Fold Validation compelete\n')
    
    return results_entry

from openpyxl import load_workbook

def save_results_to_xlsx_append(results, output_dir, filename, sheet_name='K-Fold Results'):
    """
    Appends results to an existing Excel file or creates a new file if it doesn't exist.

    Args:
        results (list or pd.DataFrame): The results data to save.
        output_dir (str): The directory where the Excel file will be saved.
        filename (str): The name of the Excel file.
        sheet_name (str): The sheet name in the Excel file. Default is 'K-Fold Results'.

    Returns:
        str: The path of the saved Excel file.
    """
    # Convert results to DataFrame if necessary
    if not isinstance(results, pd.DataFrame):
        results_df = pd.DataFrame(results)
    else:
        results_df = results

    # Rearrange columns if "Identifier" is present
    if 'Identifier' in results_df.columns:
        columns_order = ['Identifier'] + [col for col in results_df.columns if col != 'Identifier']
        results_df = results_df[columns_order]

    # Ensure output directory exists
    os.makedirs(output_dir, exist_ok=True)

    # Define the full output path
    output_path = os.path.join(output_dir, filename)

    # Append to existing Excel file or create a new one
    if os.path.exists(output_path):
        print(f"Appending data to existing file: {output_path}")
        with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # Get the existing workbook
            existing_workbook = load_workbook(output_path)

            # Check if the sheet exists
            if sheet_name in existing_workbook.sheetnames:
                # Load existing sheet and append
                start_row = existing_workbook[sheet_name].max_row
                results_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=start_row)
            else:
                # Write new sheet if not exists
                results_df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        print(f"Creating new file: {output_path}")
        results_df.to_excel(output_path, index=False, sheet_name=sheet_name)

    print(f"Results successfully saved to: {output_path}")
    return output_path

import time
import threading

def shutdown_with_countdown(countdown_seconds=30):
    """
    Initiates a shutdown countdown, allowing the user to cancel shutdown within the given time.

    Args:
        countdown_seconds (int): The number of seconds to wait before shutting down.
    """
    def cancel_shutdown():
        nonlocal shutdown_flag
        user_input = input("\nPress 'c' and Enter to cancel shutdown: ").strip().lower()
        if user_input == 'c':
            shutdown_flag = False
            print("Shutdown cancelled.")

    # Flag to determine whether to proceed with shutdown
    shutdown_flag = True

    # Start a thread to listen for user input
    input_thread = threading.Thread(target=cancel_shutdown, daemon=True)
    input_thread.start()

    # Countdown timer
    print(f"Shutdown scheduled in {countdown_seconds} seconds. Press 'c' to cancel.")
    for i in range(countdown_seconds, 0, -1):
        print(f"Time remaining: {i} seconds", end="\r")
        time.sleep(1)

    # Check the flag after countdown
    if shutdown_flag:
        print("\nShutdown proceeding...")
        os.system("shutdown /s /t 1")  # Execute shutdown command
    else:
        print("\nShutdown aborted.")

def end_program_actions(play_sound=True, shutdown=False, countdown_seconds=120):
    """
    Performs actions at the end of the program, such as playing a sound or shutting down the system.

    Args:
        play_sound (bool): If True, plays a notification sound.
        shutdown (bool): If True, initiates shutdown with a countdown.
        countdown_seconds (int): Countdown time for shutdown confirmation.
    """
    if play_sound:
        try:
            import winsound
            print("Playing notification sound...")
            winsound.Beep(1000, 500)  # Frequency: 1000Hz, Duration: 500ms
        except ImportError:
            print("winsound module not available. Skipping sound playback.")

    if shutdown:
        shutdown_with_countdown(countdown_seconds)

# %% Usage
import feature_engineering
import vce_modeling
from models import models

# %% Labels
labels = utils_feature_loading.read_labels(dataset='seed')
y = torch.tensor(np.array(labels)).view(-1)

# %% CM
# circle
all_results_original = []
for sub in range(1, 16):
    for ex in range(1, 4):
        subject_id = f"sub{sub}ex{ex}"
        print(f"Evaluating {subject_id}...")
        
        # CM
        features = utils_feature_loading.read_fcs_mat('seed', subject_id, 'pcc')
        alpha = features['alpha']
        beta = features['beta']
        gamma = features['gamma']
        
        x = np.stack((alpha, beta, gamma), axis=1)
        
        # model
        model = models.CNN_2layers_adaptive_maxpool_3()
        # evaluation; training and testing
        result_CM = cnn_validation.cnn_cross_validation(model, x, y)
        
        # Add identifier to the result
        result_CM['Identifier'] = f'sub{sub}ex{ex}'
        
        all_results_original.append(result_CM)
            
# print(f'Final Results: {results_entry}')
print('K-Fold Validation compelete\n')

# Save results to XLSX (append mode)
output_dir = os.path.join(os.getcwd(), 'Results')
filename_CM = "cnn_validation_CM_PCC.xlsx"
save_results_to_xlsx_append(all_results_original, output_dir, filename_CM)

# %% RCM
# VCE model: distance matrix & factor matrix
_, distance_matrix = feature_engineering.compute_distance_matrix('seed', {'type': '3d'})
distance_matrix = feature_engineering.normalize_matrix(distance_matrix)

factor_matrix = vce_modeling.compute_volume_conduction_factors_basic_model(
    distance_matrix, 'sigmoid', {'mu': 2.0, 'beta': 1.0})

factor_matrix_normalized = feature_engineering.normalize_matrix(factor_matrix)

# circle
all_results_recovered = []
for sub in range(1, 16):
    for ex in range(1, 4):
        subject_id = f"sub{sub}ex{ex}"
        print(f"Evaluating {subject_id}...")
        
        # CM
        features = utils_feature_loading.read_fcs_mat('seed', subject_id, 'pcc')
        alpha = features['alpha']
        beta = features['beta']
        gamma = features['gamma']
        
        x = np.stack((alpha, beta, gamma), axis=1)

        # RCM
        alpha_recovered = alpha - factor_matrix_normalized
        beta_recovered = beta - factor_matrix_normalized
        gamma_recovered = gamma - factor_matrix_normalized
        x_recovered = np.stack((alpha_recovered, beta_recovered, gamma_recovered), axis=1)
        
        # model
        model = models.CNN_2layers_adaptive_maxpool_3()
        # evaluation; training and testing
        result_RCM = cnn_validation.cnn_cross_validation(model, x_recovered, y)
        
        # Add identifier to the result
        result_RCM['Identifier'] = f'sub{sub}ex{ex}'
        
        all_results_recovered.append(result_RCM)
            
# print(f'Final Results: {results_entry}')
print('K-Fold Validation compelete\n')

# save results to XLSX (append mode)
output_dir = os.path.join(os.getcwd(), 'Results')
filename_RCM = "cnn_validation_RCM_PCC.xlsx"
save_results_to_xlsx_append(all_results_recovered, output_dir, filename_RCM)

# %% End
end_program_actions(play_sound=True, shutdown=False, countdown_seconds=120)